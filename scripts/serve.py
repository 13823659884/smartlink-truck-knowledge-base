from __future__ import annotations

import copy
import base64
import binascii
import csv
import io
import json
import mimetypes
import re
import sqlite3
import sys
import threading
import time
import uuid
from collections import OrderedDict, deque
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, urlparse

from doubao_client import (
    agent_status,
    generate_batch_diagnosis,
    generate_grounded_answer,
    generate_grounded_answer_stream,
)
from diagnosis import (
    diagnosis_summary,
    resolve_diagnostic_reply,
    select_pending_question,
    structure_fault_code,
)
from ocr_engine import ocr_status, recognize_image_bytes
from query_kb import fetch_triples, search_knowledge_base, warm_chunk_cache
from qdrant_store import semantic_search, status as qdrant_status
from source_preview import create_preview
from task_router import (
    TASK_LABELS,
    TASK_SCENES,
    classify_batch_question,
    normalize_task_type,
    rewrite_engineer_question,
)


BASE_DIR = Path(__file__).resolve().parents[1]
DATABASE_PATH = BASE_DIR / "output" / "knowledge_base.db"
WEB_DIR = BASE_DIR / "web"
CONFIG = json.loads((BASE_DIR / "config.json").read_text(encoding="utf-8"))
SOURCE_ROOT = (BASE_DIR / CONFIG["source_root"]).resolve()
PREVIEW_DIR = BASE_DIR / "output" / "previews"
BUILD_REPORT_PATH = BASE_DIR / "output" / "build_report.json"
HOST = "127.0.0.1"
PORT = 8008
RETRIEVAL_CONFIG = CONFIG.get("retrieval", {})
RETRIEVAL_BACKEND = str(RETRIEVAL_CONFIG.get("backend", "sqlite_hybrid"))
CANDIDATE_LIMIT = int(RETRIEVAL_CONFIG.get("candidate_limit", 240))
SEMANTIC_LIMIT = int(RETRIEVAL_CONFIG.get("semantic_limit", 80))
RRF_K = int(RETRIEVAL_CONFIG.get("rrf_k", 60))
CACHE_MAX_ENTRIES = int(RETRIEVAL_CONFIG.get("cache_max_entries", 256))
CACHE_TTL_SECONDS = int(RETRIEVAL_CONFIG.get("cache_ttl_seconds", 600))
PERFORMANCE_SAMPLE_SIZE = int(
    RETRIEVAL_CONFIG.get("performance_sample_size", 500)
)
PREVIEW_WARM_WORKERS = int(RETRIEVAL_CONFIG.get("preview_warm_workers", 2))

RETRIEVAL_CACHE: OrderedDict[
    tuple[object, ...], tuple[float, dict[str, object]]
] = OrderedDict()
RETRIEVAL_CACHE_LOCK = threading.Lock()
PERFORMANCE_SAMPLES: deque[dict[str, object]] = deque(
    maxlen=max(20, PERFORMANCE_SAMPLE_SIZE)
)
PERFORMANCE_LOCK = threading.Lock()
PREVIEW_EXECUTOR = ThreadPoolExecutor(
    max_workers=max(1, min(PREVIEW_WARM_WORKERS, 4)),
    thread_name_prefix="kb-preview",
)

DEFAULT_QUERY_INTENTS = (
    ("all", "智能问答", "", "自动识别问题意图并查询企业知识库", "刹车不灵敏怎么排查？", 0),
    ("vin", "VIN查询", "", "查询车辆静态字段和车型信息", "请输入17位VIN", 10),
    ("fault", "故障码查询", "修", "查询P码、SPN+FMI、故障名称与诊断指导", "例如：P312700或SPN 520216 FMI 31", 20),
    ("symptom", "症状查询", "修", "根据故障现象进行诊断和连续追问", "例如：动力不足、制动不灵敏", 30),
    ("usage", "用车知识", "用", "查询车辆操作和日常使用知识", "例如：驻车再生怎么操作？", 40),
    ("maintenance", "保养知识", "养", "查询保养周期、油液和部件维护要求", "例如：变速箱多久保养？", 50),
    ("warranty", "保用知识", "保", "查询保用标准和适用条件", "例如：制动部件是否在保？", 60),
)
INTENT_SCENES = {item[0]: item[2] for item in DEFAULT_QUERY_INTENTS}
PROFESSIONAL_INTENTS = {"vin", "fault", "symptom", "usage", "maintenance", "warranty"}
INTENT_TASK_TYPES = {
    "vin": "vin",
    "fault": "fault_code",
    "symptom": "symptom_diagnosis",
    "usage": "usage",
    "maintenance": "maintenance",
    "warranty": "warranty",
}
VALID_TASK_TYPES = {
    "vin", "fault_code", "symptom_diagnosis", "usage", "maintenance",
    "warranty", "service_technical", "drawing", "claim_case", "general",
}
PROFESSIONAL_MARKERS = (
    "车辆", "卡车", "轻卡", "重卡", "货车", "解放", "发动机", "变速箱", "离合器",
    "制动", "刹车", "故障码", "故障灯", "维修", "检修", "保养", "保用", "索赔",
    "轮胎", "底盘", "电池", "高压", "充电", "再生", "机油", "冷却液", "尿素",
    "vin", "spn", "fmi", "故障", "空压机", "驾驶室", "车桥", "仪表",
)


def question_scope(question: str, intent: str, scene: str) -> str:
    """Route vehicle-domain questions to KB-only mode and other questions to general mode."""
    normalized = str(question or "").strip().lower()
    if intent in PROFESSIONAL_INTENTS or scene in {"用", "养", "修", "保"}:
        return "professional"
    if any(marker.lower() in normalized for marker in PROFESSIONAL_MARKERS):
        return "professional"
    return "general"


def task_type_for_payload(payload: dict[str, object]) -> str:
    explicit = str(payload.get("task_type", "")).strip()
    if explicit in VALID_TASK_TYPES:
        return explicit
    return INTENT_TASK_TYPES.get(str(payload.get("intent", "")).strip(), "")


def general_retrieval_result(
    question: str, vehicle_series: str, scene: str, energy_type: str
) -> dict[str, object]:
    return {
        "question": question,
        "filters": {
            "vehicle_series": vehicle_series,
            "scene": scene,
            "energy_type": energy_type,
        },
        "answer": "",
        "sources": [],
        "triples": [],
        "retrieval": {
            "scope": "general",
            "candidate_count": 0,
            "lexical_candidate_count": 0,
            "semantic_candidate_count": 0,
            "source_count": 0,
            "triple_count": 0,
            "method": "通用问答：本轮不检索企业知识库",
            "timing_ms": {"candidate_fetch": 0, "semantic": 0, "rerank": 0, "graph": 0, "total": 0},
        },
    }


def ensure_runtime_schema() -> None:
    with sqlite3.connect(DATABASE_PATH) as connection:
        columns = {
            str(row[1])
            for row in connection.execute("PRAGMA table_info(conversations)")
        }
        if "pending_question" not in columns:
            connection.execute(
                "ALTER TABLE conversations ADD COLUMN pending_question TEXT DEFAULT ''"
            )
        if "diagnostic_topic" not in columns:
            connection.execute(
                "ALTER TABLE conversations ADD COLUMN diagnostic_topic TEXT DEFAULT ''"
            )
        feedback_columns = {
            str(row[1]) for row in connection.execute("PRAGMA table_info(feedback)")
        }
        if "comment" not in feedback_columns:
            connection.execute("ALTER TABLE feedback ADD COLUMN comment TEXT DEFAULT ''")
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS query_intents (
              id TEXT PRIMARY KEY, name TEXT NOT NULL, scene TEXT DEFAULT '',
              description TEXT DEFAULT '', example_question TEXT DEFAULT '',
              sort_order INTEGER NOT NULL DEFAULT 0,
              enabled INTEGER NOT NULL DEFAULT 1
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS vin_records (
              vin TEXT PRIMARY KEY, vehicle_type TEXT DEFAULT '',
              chassis_no TEXT DEFAULT '',
              emission_type TEXT DEFAULT '', vehicle_series TEXT DEFAULT '',
              fuel_type TEXT DEFAULT '', announcement_model TEXT DEFAULT '',
              factory_model_code TEXT DEFAULT '', rear_axle TEXT DEFAULT '',
              tire_spec TEXT DEFAULT '', engine_type TEXT DEFAULT '',
              engine_model TEXT DEFAULT '', transmission_model TEXT DEFAULT '',
              offline_time TEXT DEFAULT '', vehicle_note TEXT DEFAULT '',
              engine_name TEXT DEFAULT '', device_app_version TEXT DEFAULT '',
              mcu_version TEXT DEFAULT '', sim_match TEXT DEFAULT '',
              updated_at TEXT NOT NULL
            )
            """
        )
        vin_columns = {
            str(row[1]) for row in connection.execute("PRAGMA table_info(vin_records)")
        }
        for name in ("vehicle_type", "device_app_version", "mcu_version", "sim_match"):
            if name not in vin_columns:
                connection.execute(
                    f"ALTER TABLE vin_records ADD COLUMN {name} TEXT DEFAULT ''"
                )
        connection.executemany(
            """
            INSERT OR IGNORE INTO query_intents
            (id, name, scene, description, example_question, sort_order)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            DEFAULT_QUERY_INTENTS,
        )
        connection.commit()


def scene_for_payload(payload: dict[str, object]) -> str:
    scene = str(payload.get("scene", "")).strip()
    intent = str(payload.get("intent", "")).strip()
    mapped = INTENT_SCENES.get(intent, "")
    return mapped or scene


def vehicle_series_for_request(vehicle_id: str, explicit_series: str) -> str:
    if explicit_series or not re.fullmatch(r"[A-HJ-NPR-Z0-9]{17}", vehicle_id.upper()):
        return explicit_series
    with connect() as connection:
        row = connection.execute(
            "SELECT vehicle_series FROM vin_records WHERE vin = ?",
            (vehicle_id.upper(),),
        ).fetchone()
    return str(row["vehicle_series"] or "") if row else ""


def model_question_for_intent(question: str, intent: str) -> str:
    requirements = {
        "fault": (
            "请按故障码查询格式回答：优先给出控制器名称、P码或SPN+FMI、"
            "故障名称、故障原因、故障等级、触发条件、当前状态处置、清除条件和诊断步骤。"
        ),
        "symptom": "请按症状诊断格式回答，区分可能原因、由易到难的检查步骤，并给出下一步确认问题。",
        "usage": "请按用车知识格式回答，突出正确操作、适用条件、禁止事项和注意事项。",
        "maintenance": "请按保养知识格式回答，突出保养周期、适用车型、材料规格、操作步骤和注意事项。",
        "warranty": "请按保用知识格式回答，明确适用范围、时限/里程、判定条件、除外责任和所需凭证。",
    }
    instruction = requirements.get(intent, "")
    if not instruction:
        return question
    return (
        f"{question}\n\n【查询意图要求】{instruction}"
        "每个结论段和每个步骤末尾使用【资料1】格式标注真实依据，资料序号必须与证据列表一致。"
    )


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def index_generation() -> int:
    if BUILD_REPORT_PATH.exists():
        return BUILD_REPORT_PATH.stat().st_mtime_ns
    return DATABASE_PATH.stat().st_mtime_ns if DATABASE_PATH.exists() else 0


def cached_retrieval(
    question: str,
    vehicle_series: str,
    scene: str,
    energy_type: str,
    context: str,
    task_type_override: str = "",
) -> tuple[dict[str, object], bool]:
    if RETRIEVAL_BACKEND not in {"sqlite_hybrid", "qdrant_hybrid"}:
        raise RuntimeError(f"尚未配置检索后端：{RETRIEVAL_BACKEND}")
    key: tuple[object, ...] = (
        index_generation(),
        question,
        vehicle_series,
        scene,
        energy_type,
        context,
        task_type_override,
        CANDIDATE_LIMIT,
        SEMANTIC_LIMIT,
        RRF_K,
    )
    now = time.monotonic()
    with RETRIEVAL_CACHE_LOCK:
        cached = RETRIEVAL_CACHE.get(key)
        if cached and now - cached[0] <= CACHE_TTL_SECONDS:
            RETRIEVAL_CACHE.move_to_end(key)
            return copy.deepcopy(cached[1]), True
        if cached:
            RETRIEVAL_CACHE.pop(key, None)

    result = search_knowledge_base(
        DATABASE_PATH,
        question,
        vehicle_series=vehicle_series,
        scene=scene,
        energy_type=energy_type,
        top_k=12,
        context=context,
        candidate_limit=CANDIDATE_LIMIT,
        semantic_limit=(SEMANTIC_LIMIT if RETRIEVAL_BACKEND == "qdrant_hybrid" else 0),
        rrf_k=RRF_K,
        task_type_override=task_type_override,
    )
    with RETRIEVAL_CACHE_LOCK:
        RETRIEVAL_CACHE[key] = (now, copy.deepcopy(result))
        RETRIEVAL_CACHE.move_to_end(key)
        while len(RETRIEVAL_CACHE) > max(1, CACHE_MAX_ENTRIES):
            RETRIEVAL_CACHE.popitem(last=False)
    return result, False


def record_performance(sample: dict[str, object]) -> None:
    with PERFORMANCE_LOCK:
        PERFORMANCE_SAMPLES.append(dict(sample))


def performance_snapshot() -> dict[str, object]:
    with PERFORMANCE_LOCK:
        samples = list(PERFORMANCE_SAMPLES)

    def summarize(field: str) -> dict[str, float]:
        values = sorted(float(item.get(field, 0.0)) for item in samples)
        if not values:
            return {"average": 0.0, "p50": 0.0, "p95": 0.0, "maximum": 0.0}

        def percentile(ratio: float) -> float:
            index = min(len(values) - 1, round((len(values) - 1) * ratio))
            return round(values[index], 2)

        return {
            "average": round(sum(values) / len(values), 2),
            "p50": percentile(0.50),
            "p95": percentile(0.95),
            "maximum": round(values[-1], 2),
        }

    cache_hits = sum(1 for item in samples if item.get("cache_hit"))
    return {
        "sample_count": len(samples),
        "retrieval_ms": summarize("retrieval_ms"),
        "agent_ms": summarize("agent_ms"),
        "total_ms": summarize("total_ms"),
        "cache_hit_rate": round(cache_hits / len(samples), 4) if samples else 0.0,
        "cache_entries": len(RETRIEVAL_CACHE),
        "cache_max_entries": CACHE_MAX_ENTRIES,
        "cache_ttl_seconds": CACHE_TTL_SECONDS,
        "backend": RETRIEVAL_BACKEND,
        "candidate_limit": CANDIDATE_LIMIT,
        "semantic_limit": SEMANTIC_LIMIT,
    }


def connect() -> sqlite3.Connection:
    connection = sqlite3.connect(DATABASE_PATH)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    return connection


def resolve_source(relative_path: str) -> Path:
    target = (SOURCE_ROOT / relative_path).resolve()
    if target != SOURCE_ROOT and SOURCE_ROOT not in target.parents:
        raise ValueError("资料路径超出语料目录")
    if not target.exists() or not target.is_file():
        raise ValueError("资料文件不存在")
    return target


def warm_preview(relative_path: str, locator: str) -> None:
    try:
        create_preview(resolve_source(relative_path), locator, PREVIEW_DIR)
    except Exception:
        return


def schedule_preview_warmup(sources: list[dict[str, object]]) -> None:
    scheduled: set[tuple[str, str]] = set()
    for source in sources:
        relative_path = str(source.get("relative_path", ""))
        locator = str(source.get("source_locator", ""))
        suffix = Path(relative_path).suffix.lower()
        key = (relative_path, locator)
        if (
            not relative_path
            or key in scheduled
            or suffix
            not in {".pdf", ".pptx", ".docx", ".png", ".jpg", ".jpeg", ".webp", ".bmp"}
        ):
            continue
        scheduled.add(key)
        PREVIEW_EXECUTOR.submit(warm_preview, relative_path, locator)
        if len(scheduled) >= 4:
            break


def warm_retrieval_runtime() -> None:
    """Pay the embedding model's one-time startup cost before the first question."""
    if RETRIEVAL_BACKEND != "qdrant_hybrid":
        return
    started = time.perf_counter()
    try:
        cached_chunks = warm_chunk_cache(DATABASE_PATH)
        semantic_search("车辆故障诊断", limit=1)
        elapsed_ms = round((time.perf_counter() - started) * 1000, 2)
        print(
            f"Qdrant 语义检索已预热：{elapsed_ms} ms，内存元数据 {cached_chunks} 条",
            flush=True,
        )
    except Exception as exc:
        print(f"Qdrant 语义检索预热失败：{type(exc).__name__}: {exc}", file=sys.stderr, flush=True)


def enrich_sources(
    result: dict[str, object], *, include_images: bool = False
) -> None:
    documents: list[dict[str, object]] = []
    images: list[dict[str, object]] = []
    references: list[dict[str, object]] = []
    seen_documents: set[str] = set()
    seen_images: set[tuple[str, str]] = set()
    previewable = {".pdf", ".pptx", ".docx", ".png", ".jpg", ".jpeg", ".webp", ".bmp"}
    for source in result.get("sources", []):
        relative_path = str(source.get("relative_path", ""))
        locator = str(source.get("source_locator", ""))
        encoded_path = quote(relative_path, safe="")
        encoded_locator = quote(locator, safe="")
        document_url = f"/api/source/file?path={encoded_path}"
        suffix = Path(relative_path).suffix.lower()
        open_url = document_url
        if suffix == ".pdf":
            page_match = re.search(r"(\d+)", locator)
            if page_match:
                open_url = f"{document_url}#page={int(page_match.group(1))}"
        source["document_url"] = open_url
        source["file_url"] = document_url
        preview_url = ""
        if include_images and suffix in previewable:
            preview_url = (
                f"/api/source/preview?path={encoded_path}&locator={encoded_locator}"
            )
            source["preview_url"] = preview_url
        if relative_path not in seen_documents and len(documents) < 6:
            seen_documents.add(relative_path)
            documents.append(
                {
                    "file_name": source.get("file_name", Path(relative_path).name),
                    "relative_path": relative_path,
                    "source_locator": locator,
                    "url": open_url,
                    "file_url": document_url,
                }
            )
        if len(references) < 6:
            references.append(
                {
                    "file_name": source.get("file_name", Path(relative_path).name),
                    "source_locator": locator,
                    "excerpt": str(source.get("excerpt", ""))[:600],
                    "document_url": open_url,
                    "file_url": document_url,
                    "score": source.get("score", 0),
                }
            )
        image_key = (relative_path, locator)
        if preview_url and image_key not in seen_images and len(images) < 8:
            seen_images.add(image_key)
            images.append(
                {
                    "file_name": source.get("file_name", Path(relative_path).name),
                    "source_locator": locator,
                    "url": preview_url,
                    "document_url": open_url,
                    "file_url": document_url,
                }
            )
    result["related_documents"] = documents
    result["related_images"] = images
    result["answer_images"] = images[:4]
    result["reference_materials"] = references


def batch_fallback_pairs(
    sources: list[dict[str, object]], limit: int = 4
) -> list[dict[str, str]]:
    """Build traceable provisional pairs when the generation API is unavailable."""
    pairs: list[dict[str, str]] = []
    seen: set[str] = set()
    cause_markers = re.compile(
        r"(?:可能原因|故障原因|原因|根因)\s*[:：]?\s*([^。；\n]{4,120})"
    )
    repair_markers = re.compile(
        r"(?:处理方法|维修方法|维修措施|排查方法|处理措施|解决方法)\s*[:：]?\s*([^\n]{6,260})"
    )
    for index, source in enumerate(sources[:10], start=1):
        excerpt = re.sub(r"\s+", " ", str(source.get("excerpt", ""))).strip()
        if not excerpt:
            continue
        cause_match = cause_markers.search(excerpt)
        repair_match = repair_markers.search(excerpt)
        cause = cause_match.group(1).strip(" ：:，,。；;") if cause_match else ""
        # Fallback export must not turn an arbitrary retrieved sentence into a
        # diagnosis. Keep only evidence containing an explicit cause label.
        if not cause or not repair_match or cause.lower().startswith("ppt"):
            continue
        normalized = re.sub(r"\W+", "", cause)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        repair_plan = (
            repair_match.group(1).strip()
            if repair_match
            else (
                "1. 按该资料定位相关系统与部件；"
                "2. 对照有效版维修手册检查线路、连接器、供电及部件状态；"
                "3. 确认异常后修复或更换，并清除故障后复测。"
            )
        )
        pairs.append(
            {
                "cause": cause,
                "repair_plan": repair_plan,
                "verification": "修复后复现原工况，确认故障现象消失且相关故障码不再出现。",
                "evidence": f"资料{index}",
            }
        )
        if len(pairs) >= limit:
            break
    return pairs


class KnowledgeHandler(BaseHTTPRequestHandler):
    protocol_version = "HTTP/1.1"

    server_version = "TruckKnowledgeBase/1.0"

    def log_message(self, format: str, *args: object) -> None:
        sys.stdout.write(
            f"{self.address_string()} - [{self.log_date_time_string()}] "
            f"{format % args}\n"
        )

    def send_json(
        self, payload: object, status: HTTPStatus = HTTPStatus.OK
    ) -> None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        if isinstance(payload, dict) and isinstance(payload.get("timing"), dict):
            timing = payload["timing"]
            self.send_header(
                "Server-Timing",
                ", ".join(
                    [
                        f"retrieval;dur={float(timing.get('retrieval_ms', 0)):.2f}",
                        f"agent;dur={float(timing.get('agent_ms', 0)):.2f}",
                        f"total;dur={float(timing.get('total_ms', 0)):.2f}",
                    ]
                ),
            )
        self.end_headers()
        self.wfile.write(data)

    def start_stream(self) -> None:
        self.send_response(HTTPStatus.OK)
        self.send_header(
            "Content-Type", "application/x-ndjson; charset=utf-8"
        )
        self.send_header("Transfer-Encoding", "chunked")
        self.send_header("Cache-Control", "no-store, no-transform")
        self.send_header("X-Accel-Buffering", "no")
        self.end_headers()

    def send_stream_event(self, payload: dict[str, object]) -> None:
        data = (
            json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
            + "\n"
        ).encode("utf-8")
        self.wfile.write(f"{len(data):X}\r\n".encode("ascii"))
        self.wfile.write(data)
        self.wfile.write(b"\r\n")
        self.wfile.flush()

    def finish_stream(self) -> None:
        self.wfile.write(b"0\r\n\r\n")
        self.wfile.flush()

    def read_json(self) -> dict[str, object]:
        length = int(self.headers.get("Content-Length", "0") or "0")
        if length <= 0 or length > 12_000_000:
            raise ValueError("请求体为空或过大")
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/intents":
            with connect() as connection:
                rows = connection.execute(
                    """
                    SELECT id, name, scene, description, example_question
                    FROM query_intents WHERE enabled = 1
                    ORDER BY sort_order, name
                    """
                ).fetchall()
            self.send_json({"intents": [dict(row) for row in rows]})
            return
        if parsed.path == "/api/vin":
            vin = parse_qs(parsed.query).get("q", [""])[0].strip().upper()
            if not re.fullmatch(r"[A-HJ-NPR-Z0-9]{17}", vin):
                self.send_json(
                    {"error": "请输入正确的17位VIN（不包含I、O、Q）"},
                    HTTPStatus.BAD_REQUEST,
                )
                return
            with connect() as connection:
                row = connection.execute(
                    "SELECT * FROM vin_records WHERE vin = ?", (vin,)
                ).fetchone()
                total = connection.execute("SELECT COUNT(*) FROM vin_records").fetchone()[0]
            self.send_json(
                {
                    "found": bool(row),
                    "vin": vin,
                    "record": dict(row) if row else None,
                    "record_count": int(total),
                    "message": (
                        "已查询到车辆静态信息"
                        if row
                        else "当前尚未导入该VIN的车辆主数据，可继续按车型查询知识库"
                    ),
                }
            )
            return
        if parsed.path == "/api/history":
            params = parse_qs(parsed.query)
            conversation_id = params.get("conversation_id", [""])[0].strip()
            with connect() as connection:
                if conversation_id:
                    rows = connection.execute(
                        """
                        SELECT id, role, content, citations_json, created_at
                        FROM messages WHERE conversation_id = ? ORDER BY id
                        """,
                        (conversation_id,),
                    ).fetchall()
                    messages = []
                    for row in rows:
                        item = dict(row)
                        try:
                            item["citations"] = json.loads(item.pop("citations_json") or "[]")
                        except (TypeError, ValueError, json.JSONDecodeError):
                            item["citations"] = []
                        messages.append(item)
                    self.send_json({"conversation_id": conversation_id, "messages": messages})
                    return
                limit = max(1, min(50, int(params.get("limit", ["12"])[0])))
                rows = connection.execute(
                    """
                    SELECT c.id, c.vehicle_id, c.vehicle_series, c.scene, c.updated_at,
                      (SELECT content FROM messages m WHERE m.conversation_id = c.id
                       AND m.role = 'user' ORDER BY m.id LIMIT 1) AS question,
                      (SELECT content FROM messages m WHERE m.conversation_id = c.id
                       AND m.role = 'assistant' ORDER BY m.id DESC LIMIT 1) AS answer
                    FROM conversations c
                    WHERE EXISTS (SELECT 1 FROM messages m WHERE m.conversation_id = c.id)
                    ORDER BY c.updated_at DESC LIMIT ?
                    """,
                    (limit,),
                ).fetchall()
            self.send_json({"history": [dict(row) for row in rows]})
            return
        if parsed.path == "/api/health":
            self.send_json(
                {
                    "status": "ok",
                    "database_exists": DATABASE_PATH.exists(),
                    "agent": agent_status(),
                    "ocr": ocr_status(),
                    "retrieval": {
                        "backend": RETRIEVAL_BACKEND,
                        "candidate_limit": CANDIDATE_LIMIT,
                        "semantic_limit": SEMANTIC_LIMIT,
                        "cache_entries": len(RETRIEVAL_CACHE),
                        "qdrant": qdrant_status(),
                    },
                    "time": utc_now(),
                }
            )
            return
        if parsed.path == "/api/agent/config":
            self.send_json(agent_status())
            return
        if parsed.path == "/api/performance":
            self.send_json(performance_snapshot())
            return
        if parsed.path == "/api/quality":
            report: dict[str, object] = {}
            if BUILD_REPORT_PATH.exists():
                try:
                    report = json.loads(
                        BUILD_REPORT_PATH.read_text(encoding="utf-8")
                    )
                except (OSError, ValueError, json.JSONDecodeError):
                    report = {}
            with connect() as connection:
                statuses = {
                    str(row["status"]): int(row["documents"])
                    for row in connection.execute(
                        """
                        SELECT status, COUNT(*) AS documents
                        FROM documents
                        GROUP BY status
                        ORDER BY status
                        """
                    )
                }
                pending = [
                    dict(row)
                    for row in connection.execute(
                        """
                        SELECT relative_path, file_name, status, error_message
                        FROM documents
                        WHERE status IN ('needs_ocr', 'error')
                        ORDER BY relative_path
                        LIMIT 100
                        """
                    )
                ]
            self.send_json(
                {
                    "valid": bool(report.get("valid", False)),
                    "finished_at": report.get("finished_at", ""),
                    "documents": report.get("documents", 0),
                    "chunks": report.get("chunks", 0),
                    "statuses": statuses,
                    "ocr": report.get("ocr", ocr_status()),
                    "pending_documents": pending,
                }
            )
            return
        if parsed.path == "/api/fault-code":
            query = parse_qs(parsed.query).get("q", [""])[0].strip()
            if not query:
                self.send_json(
                    {"error": "请输入故障码或故障描述"},
                    HTTPStatus.BAD_REQUEST,
                )
                return
            with connect() as connection:
                triples, method = fetch_triples(connection, query, limit=200)
            retrieval = search_knowledge_base(
                DATABASE_PATH,
                query,
                scene="修",
                top_k=8,
                candidate_limit=CANDIDATE_LIMIT,
                semantic_limit=SEMANTIC_LIMIT,
                rrf_k=RRF_K,
            )
            payload = structure_fault_code(
                query,
                triples,
                method,
                sources=retrieval.get("sources", []),
            )
            payload["answer"] = retrieval.get("answer", "")
            payload["sources"] = retrieval.get("sources", [])
            payload["retrieval"] = retrieval.get("retrieval", {})
            self.send_json(payload)
            return
        if parsed.path == "/api/source/file":
            try:
                params = parse_qs(parsed.query)
                source_path = resolve_source(params.get("path", [""])[0])
                self.send_local_file(source_path, inline=True)
            except ValueError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return
        if parsed.path == "/api/source/preview":
            try:
                params = parse_qs(parsed.query)
                source_path = resolve_source(params.get("path", [""])[0])
                locator = params.get("locator", [""])[0]
                preview = create_preview(source_path, locator, PREVIEW_DIR)
                if preview is None:
                    self.send_json(
                        {"error": "该格式暂不支持图片预览"},
                        HTTPStatus.UNSUPPORTED_MEDIA_TYPE,
                    )
                    return
                self.send_local_file(preview, inline=True)
            except ValueError as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            except Exception as exc:
                self.send_json(
                    {"error": f"预览生成失败：{type(exc).__name__}: {exc}"},
                    HTTPStatus.INTERNAL_SERVER_ERROR,
                )
            return
        if parsed.path == "/api/stats":
            with connect() as connection:
                payload = {
                    "documents": connection.execute(
                        "SELECT COUNT(*) FROM documents"
                    ).fetchone()[0],
                    "active_documents": connection.execute(
                        "SELECT COUNT(*) FROM documents WHERE enabled = 1"
                    ).fetchone()[0],
                    "chunks": connection.execute(
                        "SELECT COUNT(*) FROM chunks"
                    ).fetchone()[0],
                    "entities": connection.execute(
                        "SELECT COUNT(*) FROM entities"
                    ).fetchone()[0],
                    "triples": connection.execute(
                        "SELECT COUNT(*) FROM triples"
                    ).fetchone()[0],
                    "knowledge_bases": connection.execute(
                        "SELECT COUNT(*) FROM knowledge_bases"
                    ).fetchone()[0],
                    "scenes": [
                        dict(row)
                        for row in connection.execute(
                            """
                            SELECT scene, COUNT(*) AS documents
                            FROM documents
                            WHERE enabled = 1
                            GROUP BY scene
                            ORDER BY scene
                            """
                        )
                    ],
                }
            self.send_json(payload)
            return
        if parsed.path == "/api/documents":
            params = parse_qs(parsed.query)
            scene = params.get("scene", [""])[0]
            status = params.get("status", [""])[0]
            where = ["1 = 1"]
            values: list[object] = []
            if scene:
                where.append("scene = ?")
                values.append(scene)
            if status:
                where.append("status = ?")
                values.append(status)
            with connect() as connection:
                rows = connection.execute(
                    f"""
                    SELECT relative_path, file_name, extension, scene,
                           vehicle_tags, energy_tags, status, enabled,
                           chunk_count, version, effective_date, error_message
                    FROM documents
                    WHERE {" AND ".join(where)}
                    ORDER BY scene, relative_path
                    LIMIT 500
                    """,
                    values,
                ).fetchall()
            self.send_json({"documents": [dict(row) for row in rows]})
            return
        if parsed.path == "/api/triples":
            params = parse_qs(parsed.query)
            query = params.get("q", [""])[0].strip()
            predicate = params.get("predicate", [""])[0].strip()
            if query and not predicate:
                with connect() as connection:
                    triples, method = fetch_triples(connection, query, limit=200)
                self.send_json({"triples": triples, "method": method})
                return
            where = ["1 = 1"]
            values: list[object] = []
            if query:
                where.append("(subject LIKE ? OR object LIKE ?)")
                values.extend([f"%{query}%", f"%{query}%"])
            if predicate:
                where.append("predicate = ?")
                values.append(predicate)
            with connect() as connection:
                rows = connection.execute(
                    f"""
                    SELECT subject, subject_label, predicate, object,
                           object_label, source_path, source_locator,
                           confidence, review_status
                    FROM triples
                    WHERE {" AND ".join(where)}
                    LIMIT 200
                    """,
                    values,
                ).fetchall()
            self.send_json({"triples": [dict(row) for row in rows]})
            return
        if parsed.path == "/api/feedback/export.csv":
            with connect() as connection:
                rows = connection.execute(
                    """
                    SELECT account, vehicle_id, question, answer, rating,
                           comment, vehicle_series, scene, created_at
                    FROM feedback
                    ORDER BY created_at DESC
                    """
                ).fetchall()
            buffer = io.StringIO()
            writer = csv.writer(buffer)
            writer.writerow(
                [
                    "账户名",
                    "车牌号/VIN",
                    "提问问题",
                    "助手答复内容",
                    "赞踩状态",
                    "纠偏意见",
                    "车系",
                    "场景",
                    "记录时间",
                ]
            )
            writer.writerows([tuple(row) for row in rows])
            data = ("\ufeff" + buffer.getvalue()).encode("utf-8")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/csv; charset=utf-8")
            self.send_header(
                "Content-Disposition",
                'attachment; filename="qa-feedback.csv"',
            )
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return
        self.serve_static(parsed.path)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        try:
            payload = self.read_json()
            if parsed.path == "/api/batch/import":
                self.handle_batch_import(payload)
                return
            if parsed.path == "/api/batch/diagnose":
                self.handle_batch_diagnose(payload)
                return
            if parsed.path == "/api/batch/export":
                self.handle_batch_export(payload)
                return
            if parsed.path == "/api/search/stream":
                self.handle_search_stream(payload)
                return
            if parsed.path == "/api/search":
                self.handle_search(payload, force_agent=False)
                return
            if parsed.path == "/api/agent/chat":
                self.handle_search(payload, force_agent=True)
                return
            if parsed.path == "/api/feedback":
                self.handle_feedback(payload)
                return
            if parsed.path == "/api/intents":
                self.handle_intent(payload)
                return
            if parsed.path == "/api/vin":
                self.handle_vin_record(payload)
                return
            if parsed.path == "/api/image/recognize":
                self.handle_image_recognition(payload)
                return
            self.send_json(
                {"error": "接口不存在"}, HTTPStatus.NOT_FOUND
            )
        except ValueError as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
        except Exception as exc:
            self.send_json(
                {"error": f"{type(exc).__name__}: {exc}"},
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )

    def handle_batch_import(self, payload: dict[str, object]) -> None:
        encoded = str(payload.get("file_base64", "")).strip()
        if encoded.startswith("data:"):
            encoded = encoded.split(",", 1)[-1]
        if not encoded:
            raise ValueError("请选择需要导入的Excel文件")
        try:
            workbook_bytes = base64.b64decode(encoded, validate=True)
        except (binascii.Error, ValueError) as exc:
            raise ValueError("Excel文件内容无效") from exc
        if len(workbook_bytes) > 10_000_000:
            raise ValueError("单个Excel文件不能超过10MB")
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(
                io.BytesIO(workbook_bytes), read_only=True, data_only=True
            )
        except Exception as exc:
            raise ValueError(f"无法读取Excel文件：{exc}") from exc

        series_aliases = ("车系", "车型", "车辆系列", "车系名称")
        # Only consume a deliberately supplied question category. Columns such
        # as “咨询类别2” describe how a call was handled (e.g. address-book
        # transfer), not necessarily the semantic type of the vehicle question.
        category_aliases = ("问题分类", "问题类别", "类别", "分类", "任务类型")
        symptom_aliases = (
            "故障现象", "客户询问问题", "车辆问题", "咨询问题", "问题",
            "故障详情", "客户问题",
        )
        default_series = str(payload.get("default_vehicle_series", "JH6")).strip() or "JH6"
        imported: list[dict[str, object]] = []
        skipped = 0
        sheets: list[str] = []
        for worksheet in workbook.worksheets:
            sheets.append(worksheet.title)
            rows = worksheet.iter_rows(values_only=True)
            try:
                header_values = next(rows)
            except StopIteration:
                continue
            headers = [str(value or "").strip() for value in header_values]

            def column_index(aliases: tuple[str, ...]) -> int | None:
                for alias in aliases:
                    if alias in headers:
                        return headers.index(alias)
                for index, header in enumerate(headers):
                    if any(alias in header for alias in aliases):
                        return index
                return None

            symptom_index = column_index(symptom_aliases)
            series_index = column_index(series_aliases)
            category_index = next(
                (headers.index(alias) for alias in category_aliases if alias in headers),
                None,
            )
            if symptom_index is None:
                continue
            for excel_row, values in enumerate(rows, start=2):
                symptom = (
                    str(values[symptom_index] or "").strip()
                    if symptom_index < len(values)
                    else ""
                )
                if not symptom:
                    skipped += 1
                    continue
                vehicle_series = default_series
                if series_index is not None and series_index < len(values):
                    vehicle_series = str(values[series_index] or "").strip() or default_series
                explicit_category = ""
                if category_index is not None and category_index < len(values):
                    explicit_category = str(values[category_index] or "").strip()
                classification = classify_batch_question(symptom, explicit_category)
                rewritten = rewrite_engineer_question(symptom)
                imported.append(
                    {
                        "id": uuid.uuid4().hex,
                        "sheet": worksheet.title,
                        "source_row": excel_row,
                        "vehicle_series": vehicle_series,
                        "symptom": symptom,
                        "original_question": rewritten["original_question"],
                        "engineer_question": rewritten["engineer_question"],
                        **classification,
                    }
                )
                if len(imported) >= 2000:
                    raise ValueError("单次最多导入2000条问题")
        if not imported:
            raise ValueError(
                "未找到问题列，请使用“故障现象”或“客户询问问题”作为列名"
            )
        self.send_json(
            {
                "rows": imported,
                "count": len(imported),
                "skipped": skipped,
                "sheets": sheets,
                "default_vehicle_series": default_series,
            }
        )

    def handle_batch_diagnose(self, payload: dict[str, object]) -> None:
        symptom = str(payload.get("symptom", "")).strip()
        engineer_question = str(payload.get("engineer_question", "")).strip()
        vehicle_series = str(payload.get("vehicle_series", "JH6")).strip() or "JH6"
        answer_mode = "deep" if str(payload.get("answer_mode", "fast")) == "deep" else "fast"
        requested_task = normalize_task_type(str(payload.get("task_type", "")))
        rewritten = rewrite_engineer_question(engineer_question or symptom)
        engineer_question = rewritten["engineer_question"]
        classification = classify_batch_question(engineer_question, requested_task)
        task_type = str(classification["task_type"])
        task_label = str(classification["task_label"])
        scene = str(classification["scene"])
        if not symptom:
            raise ValueError("故障现象不能为空")
        if len(symptom) > 2000:
            raise ValueError("单条故障现象不能超过2000字")
        started = time.perf_counter()
        result, cache_hit = cached_retrieval(
            engineer_question,
            vehicle_series,
            scene,
            "",
            "",
            task_type,
        )
        agent_result = generate_batch_diagnosis(
            symptom=engineer_question,
            sources=result.get("sources", []),
            triples=result.get("triples", []),
            vehicle_series=vehicle_series,
            task_type=task_type,
            task_label=task_label,
            mode=answer_mode,
        )
        if not agent_result.get("ok"):
            fallback_pairs = batch_fallback_pairs(result.get("sources", []))
            if not fallback_pairs:
                self.send_json(
                    {"error": agent_result.get("error", "批量诊断失败")},
                    HTTPStatus.BAD_GATEWAY,
                )
                return
            agent_result = {
                **agent_result,
                "ok": True,
                "fallback": True,
                "summary": (
                    "模型接口当前不可用，以下为知识库检索资料的临时整理结果；"
                    "请由维修人员结合有效版手册复核。"
                ),
                "pairs": fallback_pairs,
                "safety_notice": "安全关键系统请由合格维修人员按有效版维修手册操作。",
            }
        enrich_sources(result, include_images=False)
        references = []
        for index, source in enumerate(result.get("sources", [])[:10], start=1):
            references.append(
                {
                    "index": index,
                    "file_name": source.get("file_name", ""),
                    "source_locator": source.get("source_locator", ""),
                    "document_url": source.get("document_url", source.get("file_url", "")),
                }
            )
        self.send_json(
            {
                "vehicle_series": vehicle_series,
                "symptom": symptom,
                "original_question": symptom,
                "engineer_question": engineer_question,
                "classification": classification,
                "summary": agent_result.get("summary", ""),
                "pairs": agent_result.get("pairs", []),
                "safety_notice": agent_result.get("safety_notice", ""),
                "references": references,
                "agent": {
                    key: value for key, value in agent_result.items()
                    if key not in {"summary", "pairs", "safety_notice"}
                },
                "retrieval": {
                    "cache_hit": cache_hit,
                    "source_count": len(result.get("sources", [])),
                    "backend": RETRIEVAL_BACKEND,
                },
                "timing_ms": round((time.perf_counter() - started) * 1000, 2),
            }
        )

    def handle_batch_export(self, payload: dict[str, object]) -> None:
        rows = payload.get("rows", [])
        if not isinstance(rows, list) or not rows:
            raise ValueError("没有可导出的批量诊断结果")
        if len(rows) > 20_000:
            raise ValueError("单次最多导出20000行")
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "批量诊断结果"
        headers = [
            "序号", "原始工作表", "原始行号", "车系", "问题分类", "客服原始问题", "工程师问题", "原因/结论",
            "对应维修方案", "验证方法", "总体判断", "安全提示", "参考资料",
            "处理状态", "错误信息", "回答模型",
        ]
        worksheet.append(headers)
        for index, row in enumerate(rows, start=1):
            if not isinstance(row, dict):
                continue
            worksheet.append(
                [
                    index,
                    str(row.get("sheet", "")),
                    row.get("source_row", ""),
                    str(row.get("vehicle_series", "")),
                    str(row.get("task_label", "")),
                    str(row.get("original_question", row.get("symptom", ""))),
                    str(row.get("engineer_question", "")),
                    str(row.get("cause", "")),
                    str(row.get("repair_plan", "")),
                    str(row.get("verification", "")),
                    str(row.get("summary", "")),
                    str(row.get("safety_notice", "")),
                    str(row.get("references", "")),
                    str(row.get("status", "")),
                    str(row.get("error", "")),
                    str(row.get("model", "")),
                ]
            )
        header_fill = PatternFill("solid", fgColor="0B5FA5")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2EC")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)
        widths = [8, 14, 10, 10, 14, 34, 34, 34, 58, 30, 34, 30, 34, 12, 28, 24]
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = width
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=thin)
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.row_dimensions[1].height = 28
        buffer = io.BytesIO()
        workbook.save(buffer)
        data = buffer.getvalue()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.send_response(HTTPStatus.OK)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header(
            "Content-Disposition",
            f"attachment; filename=batch-diagnosis-{timestamp}.xlsx",
        )
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(data)

    def handle_search(
        self, payload: dict[str, object], force_agent: bool = False
    ) -> None:
        request_started = time.perf_counter()
        question = str(payload.get("question", "")).strip()
        if not question:
            raise ValueError("请输入问题")
        image_ocr_text = str(payload.get("image_ocr_text", "")).strip()[:4000]
        conversation_id = str(
            payload.get("conversation_id") or uuid.uuid4().hex
        )
        with connect() as state_connection:
            state_row = state_connection.execute(
                """
                SELECT pending_question, diagnostic_topic
                FROM conversations
                WHERE id = ?
                """,
                (conversation_id,),
            ).fetchone()
        reply_context = resolve_diagnostic_reply(
            question,
            pending_question=(
                str(state_row["pending_question"] or "") if state_row else ""
            ),
            diagnostic_topic=(
                str(state_row["diagnostic_topic"] or "") if state_row else ""
            ),
            explicit_prompt=str(payload.get("reply_to_question", "")),
        )
        effective_question = str(reply_context["effective_question"])
        if image_ocr_text:
            effective_question += (
                f"\n\n用户上传图片的OCR识别内容：\n{image_ocr_text}"
            )
        answered_prompt = str(reply_context.get("answered_prompt", ""))
        diagnostic_topic = str(reply_context.get("topic", question))
        account = str(payload.get("account", "")).strip()
        vehicle_id = str(payload.get("vehicle_id", "")).strip()
        vehicle_series = vehicle_series_for_request(
            vehicle_id, str(payload.get("vehicle_series", "")).strip()
        )
        scene = scene_for_payload(payload)
        energy_type = str(payload.get("energy_type", "")).strip()
        scope = question_scope(effective_question, str(payload.get("intent", "")).strip(), scene)
        now = utc_now()

        persistence_started = time.perf_counter()
        with connect() as connection:
            connection.execute(
                """
                INSERT INTO conversations
                (id, account, vehicle_id, vehicle_series, scene, created_at,
                 updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(id) DO UPDATE SET
                  account = excluded.account,
                  vehicle_id = excluded.vehicle_id,
                  vehicle_series = excluded.vehicle_series,
                  scene = excluded.scene,
                  updated_at = excluded.updated_at
                """,
                (
                    conversation_id,
                    account,
                    vehicle_id,
                    vehicle_series,
                    scene,
                    now,
                    now,
                ),
            )
            history = connection.execute(
                """
                SELECT role, content
                FROM messages
                WHERE conversation_id = ?
                ORDER BY id DESC
                LIMIT 10
                """,
                (conversation_id,),
            ).fetchall()
            context = "\n".join(
                f"{row['role']}：{row['content']}"
                for row in reversed(history)
            )
            connection.execute(
                """
                INSERT INTO messages(conversation_id, role, content, created_at)
                VALUES (?, 'user', ?, ?)
                """,
                (
                    conversation_id,
                    question
                    + ("\n[已附图片识别文字]" if image_ocr_text else ""),
                    now,
                ),
            )
            connection.commit()
        persistence_ms = round(
            (time.perf_counter() - persistence_started) * 1000, 2
        )

        retrieval_started = time.perf_counter()
        if scope == "professional":
            result, cache_hit = cached_retrieval(
                effective_question,
                vehicle_series,
                scene,
                energy_type,
                context,
                task_type_for_payload(payload),
            )
        else:
            result, cache_hit = (
                general_retrieval_result(effective_question, vehicle_series, scene, energy_type),
                False,
            )
        retrieval_ms = round(
            (time.perf_counter() - retrieval_started) * 1000, 2
        )
        result["retrieval"]["cache_hit"] = cache_hit
        result["retrieval"]["backend"] = RETRIEVAL_BACKEND
        result["retrieval"]["scope"] = scope
        use_agent = force_agent or bool(payload.get("use_agent", False)) or scope == "general"
        answer_mode = (
            "fast" if str(payload.get("answer_mode", "fast")) == "fast" else "deep"
        )
        agent_started = time.perf_counter()
        if use_agent:
            agent_result = generate_grounded_answer(
                question=model_question_for_intent(
                    effective_question, str(payload.get("intent", "")).strip()
                ),
                sources=result["sources"],
                triples=result["triples"],
                history=context,
                vehicle_series=vehicle_series,
                scene=scene,
                mode=answer_mode,
                knowledge_only=scope == "professional",
            )
            result["agent"] = {
                key: value
                for key, value in agent_result.items()
                if key not in {
                    "answer",
                    "related_questions",
                    "solution_steps",
                    "safety_notice",
                }
            }
            if agent_result.get("ok"):
                result["answer"] = agent_result["answer"]
                result["related_questions"] = agent_result[
                    "related_questions"
                ]
                result["solution_steps"] = agent_result["solution_steps"]
                result["safety_notice"] = agent_result["safety_notice"]
            else:
                fallback_answer = result["answer"]
                result["answer"] = (
                    f"通用智能体当前不可用：{agent_result.get('error', '调用失败')}。"
                    if scope == "general"
                    else (
                        f"智能体当前不可用：{agent_result.get('error', '调用失败')}。\n"
                        "以下内容仅为知识库自动检索摘要，可能包含其他车型或相邻场景资料，"
                        "不代表针对当前车辆的最终维修结论。\n\n"
                        f"{fallback_answer}"
                    )
                )
        else:
            result["agent"] = {
                **agent_status(),
                "ok": False,
                "skipped": True,
            }
        agent_ms = round((time.perf_counter() - agent_started) * 1000, 2)
        include_images = bool(payload.get("include_images", False))
        enrich_sources(result, include_images=include_images)
        result["diagnosis"] = diagnosis_summary(
            effective_question,
            sources=result.get("sources", []),
            triples=result.get("triples", []),
            vehicle_series=vehicle_series,
            answered_prompt=answered_prompt,
            user_reply=question,
        )
        if scope == "general":
            result["diagnosis"] = {
                "pending_question": "",
                "reply_options": [],
                "context_applied": bool(reply_context.get("applied")),
                "scope": "general",
            }
        pending_question = (
            ""
            if scope == "general"
            else select_pending_question(
                result["diagnosis"],
                related_questions=result.get("related_questions", []),
                answered_prompt=answered_prompt,
            )
        )
        yes_no_markers = ("是否", "有没有", "能否", "亮", "正常", "一致")
        result["diagnosis"]["pending_question"] = pending_question
        result["diagnosis"]["reply_options"] = (
            ["是", "否", "不确定"]
            if pending_question
            and any(marker in pending_question for marker in yes_no_markers)
            else []
        )
        result["diagnosis"]["context_applied"] = bool(
            reply_context.get("applied")
        )
        result["conversation_context"] = {
            "continued": bool(reply_context.get("applied")),
            "answered_prompt": answered_prompt,
            "pending_question": pending_question,
            "topic": diagnostic_topic,
        }
        if image_ocr_text:
            result["image_recognition"] = {
                "used": True,
                "text": image_ocr_text,
                "file_name": str(payload.get("image_name", "图片")),
            }
        if include_images:
            schedule_preview_warmup(result.get("sources", []))
        response_persistence_started = time.perf_counter()
        with connect() as connection:
            history_content = str(result["answer"])
            if pending_question:
                history_content += f"\n\n【待确认问题】{pending_question}"
            cursor = connection.execute(
                """
                INSERT INTO messages
                (conversation_id, role, content, citations_json, created_at)
                VALUES (?, 'assistant', ?, ?, ?)
                """,
                (
                    conversation_id,
                    history_content,
                    json.dumps(result["sources"], ensure_ascii=False),
                    utc_now(),
                ),
            )
            connection.execute(
                """
                UPDATE conversations
                SET pending_question = ?, diagnostic_topic = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    pending_question,
                    diagnostic_topic,
                    utc_now(),
                    conversation_id,
                ),
            )
            connection.commit()
            message_id = cursor.lastrowid
        persistence_ms += round(
            (time.perf_counter() - response_persistence_started) * 1000, 2
        )
        result["conversation_id"] = conversation_id
        result["message_id"] = message_id
        total_ms = round((time.perf_counter() - request_started) * 1000, 2)
        result["timing"] = {
            "retrieval_ms": retrieval_ms,
            "agent_ms": agent_ms,
            "persistence_ms": round(persistence_ms, 2),
            "total_ms": total_ms,
            "cache_hit": cache_hit,
        }
        record_performance(
            {
                "retrieval_ms": retrieval_ms,
                "agent_ms": agent_ms,
                "total_ms": total_ms,
                "cache_hit": cache_hit,
                "use_agent": use_agent,
                "time": utc_now(),
            }
        )
        self.send_json(result)

    def handle_search_stream(self, payload: dict[str, object]) -> None:
        request_started = time.perf_counter()
        question = str(payload.get("question", "")).strip()
        if not question:
            raise ValueError("请输入问题")

        self.start_stream()
        try:
            self.send_stream_event(
                {"type": "status", "text": "正在分析问题类型"}
            )
            image_ocr_text = str(payload.get("image_ocr_text", "")).strip()[:4000]
            conversation_id = str(
                payload.get("conversation_id") or uuid.uuid4().hex
            )
            with connect() as state_connection:
                state_row = state_connection.execute(
                    """
                    SELECT pending_question, diagnostic_topic
                    FROM conversations WHERE id = ?
                    """,
                    (conversation_id,),
                ).fetchone()
            reply_context = resolve_diagnostic_reply(
                question,
                pending_question=(
                    str(state_row["pending_question"] or "")
                    if state_row
                    else ""
                ),
                diagnostic_topic=(
                    str(state_row["diagnostic_topic"] or "")
                    if state_row
                    else ""
                ),
                explicit_prompt=str(payload.get("reply_to_question", "")),
            )
            effective_question = str(reply_context["effective_question"])
            if image_ocr_text:
                effective_question += (
                    f"\n\n用户上传图片的OCR识别内容：\n{image_ocr_text}"
                )
            answered_prompt = str(reply_context.get("answered_prompt", ""))
            diagnostic_topic = str(reply_context.get("topic", question))
            account = str(payload.get("account", "")).strip()
            vehicle_id = str(payload.get("vehicle_id", "")).strip()
            vehicle_series = vehicle_series_for_request(
                vehicle_id, str(payload.get("vehicle_series", "")).strip()
            )
            scene = scene_for_payload(payload)
            energy_type = str(payload.get("energy_type", "")).strip()
            scope = question_scope(effective_question, str(payload.get("intent", "")).strip(), scene)
            answer_mode = (
                "fast"
                if str(payload.get("answer_mode", "fast")) == "fast"
                else "deep"
            )
            now = utc_now()

            persistence_started = time.perf_counter()
            with connect() as connection:
                connection.execute(
                    """
                    INSERT INTO conversations
                    (id, account, vehicle_id, vehicle_series, scene, created_at,
                     updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(id) DO UPDATE SET
                      account = excluded.account,
                      vehicle_id = excluded.vehicle_id,
                      vehicle_series = excluded.vehicle_series,
                      scene = excluded.scene,
                      updated_at = excluded.updated_at
                    """,
                    (
                        conversation_id,
                        account,
                        vehicle_id,
                        vehicle_series,
                        scene,
                        now,
                        now,
                    ),
                )
                history = connection.execute(
                    """
                    SELECT role, content FROM messages
                    WHERE conversation_id = ?
                    ORDER BY id DESC LIMIT 10
                    """,
                    (conversation_id,),
                ).fetchall()
                context = "\n".join(
                    f"{row['role']}：{row['content']}"
                    for row in reversed(history)
                )
                connection.execute(
                    """
                    INSERT INTO messages(conversation_id, role, content, created_at)
                    VALUES (?, 'user', ?, ?)
                    """,
                    (
                        conversation_id,
                        question
                        + ("\n[已附图片识别文字]" if image_ocr_text else ""),
                        now,
                    ),
                )
                connection.commit()
            persistence_ms = round(
                (time.perf_counter() - persistence_started) * 1000, 2
            )

            retrieval_started = time.perf_counter()
            if scope == "professional":
                result, cache_hit = cached_retrieval(
                    effective_question,
                    vehicle_series,
                    scene,
                    energy_type,
                    context,
                    task_type_for_payload(payload),
                )
            else:
                result, cache_hit = (
                    general_retrieval_result(effective_question, vehicle_series, scene, energy_type),
                    False,
                )
            retrieval_ms = round(
                (time.perf_counter() - retrieval_started) * 1000, 2
            )
            result["retrieval"]["cache_hit"] = cache_hit
            result["retrieval"]["backend"] = RETRIEVAL_BACKEND
            result["retrieval"]["scope"] = scope
            self.send_stream_event(
                {
                    "type": "status",
                    "text": "正在生成通用回答" if scope == "general" else "正在检索企业知识库并生成专业回答",
                }
            )
            self.send_stream_event(
                {
                    "type": "meta",
                    "conversation_id": conversation_id,
                    "answer_mode": answer_mode,
                    "source_count": len(result.get("sources", [])),
                    "retrieval_ms": retrieval_ms,
                    "retrieval": result.get("retrieval", {}),
                    "text": (
                        "正在生成通用回答"
                        if scope == "general"
                        else f"已找到 {len(result.get('sources', []))} 条相关资料，正在生成专业回答"
                    ),
                }
            )

            agent_started = time.perf_counter()
            agent_result: dict[str, Any] = {}
            raw_content = ""
            visible_answer = ""
            first_token_ms = 0.0
            # Fast mode is intentionally retried once before falling back. Ark's
            # streaming connection can occasionally be reset before the first
            # token; treating that transient as an unavailable model produces a
            # misleading message in the mini program.
            mode_attempts = (
                [("fast", 1), ("fast", 2), ("deep", 1)]
                if answer_mode == "fast"
                else [("deep", 1), ("fast", 1), ("fast", 2)]
            )
            for attempt_index, (active_mode, attempt) in enumerate(mode_attempts):
                raw_content = ""
                visible_answer = ""
                for event in generate_grounded_answer_stream(
                    question=model_question_for_intent(
                        effective_question, str(payload.get("intent", "")).strip()
                    ),
                    sources=result["sources"],
                    triples=result["triples"],
                    history=context,
                    vehicle_series=vehicle_series,
                    scene=scene,
                    mode=active_mode,
                    knowledge_only=scope == "professional",
                    timeout=25 if active_mode == "deep" else 8,
                ):
                    if event.get("type") == "delta":
                        raw_content += str(event.get("text", ""))
                        candidate = raw_content
                        if "【回答】" in candidate:
                            candidate = candidate.split("【回答】", 1)[1]
                        elif candidate.lstrip().startswith("【"):
                            candidate = ""
                        candidate = candidate.lstrip("\n")
                        if candidate.startswith(visible_answer):
                            delta = candidate[len(visible_answer) :]
                            if delta:
                                visible_answer = candidate
                                if not first_token_ms:
                                    first_token_ms = round(
                                        (time.perf_counter() - request_started) * 1000,
                                        2,
                                    )
                                self.send_stream_event(
                                    {"type": "delta", "text": delta}
                                )
                    elif event.get("type") == "done":
                        agent_result = dict(event)
                if agent_result.get("ok"):
                    answer_mode = active_mode
                    if not visible_answer and agent_result.get("answer"):
                        visible_answer = str(agent_result["answer"])
                        if not first_token_ms:
                            first_token_ms = round(
                                (time.perf_counter() - request_started) * 1000,
                                2,
                            )
                        self.send_stream_event(
                            {"type": "delta", "text": visible_answer}
                        )
                    break
                if not visible_answer:
                    print(
                        f"[{active_mode}-model] "
                        f"attempt={attempt} error={agent_result.get('error', 'unknown')}",
                        file=sys.stderr,
                        flush=True,
                    )
                    next_mode = (
                        mode_attempts[attempt_index + 1][0]
                        if attempt_index + 1 < len(mode_attempts)
                        else ""
                    )
                    if next_mode == active_mode:
                        self.send_stream_event(
                            {
                                "type": "status",
                                "text": (
                                    "快速模型连接波动，正在自动重试"
                                    if active_mode == "fast"
                                    else "深度模型连接波动，正在自动重试"
                                ),
                            }
                        )
                        continue
                    if next_mode:
                        self.send_stream_event(
                            {
                                "type": "mode_fallback",
                                "answer_mode": next_mode,
                                "text": (
                                    "深度模型响应超时，已切换快速模型继续回答"
                                    if active_mode == "deep"
                                    else "快速通道连续连接失败，正在切换深度模式"
                                ),
                            }
                        )
                        continue
                break
            agent_ms = round((time.perf_counter() - agent_started) * 1000, 2)

            result["agent"] = {
                key: value
                for key, value in agent_result.items()
                if key
                not in {
                    "type",
                    "answer",
                    "related_questions",
                    "solution_steps",
                    "safety_notice",
                }
            }
            result["answer_mode"] = answer_mode
            if agent_result.get("ok"):
                result["answer"] = str(agent_result.get("answer", ""))
                result["related_questions"] = list(
                    agent_result.get("related_questions", [])
                )
                result["solution_steps"] = list(
                    agent_result.get("solution_steps", [])
                )
                result["safety_notice"] = str(
                    agent_result.get("safety_notice", "")
                )
            else:
                fallback_answer = str(result["answer"])
                result["answer"] = (
                    (
                        f"通用智能体当前不可用：{agent_result.get('error', '调用失败')}。"
                        if scope == "general"
                        else (
                            f"智能体当前不可用：{agent_result.get('error', '调用失败')}。\n\n"
                            f"{fallback_answer}"
                        )
                    )
                )

            include_images = bool(payload.get("include_images", False))
            enrich_sources(result, include_images=include_images)
            result["diagnosis"] = diagnosis_summary(
                effective_question,
                sources=result.get("sources", []),
                triples=result.get("triples", []),
                vehicle_series=vehicle_series,
                answered_prompt=answered_prompt,
                user_reply=question,
            )
            if scope == "general":
                result["diagnosis"] = {
                    "pending_question": "",
                    "reply_options": [],
                    "context_applied": bool(reply_context.get("applied")),
                    "scope": "general",
                }
            pending_question = (
                ""
                if scope == "general"
                else select_pending_question(
                    result["diagnosis"],
                    related_questions=result.get("related_questions", []),
                    answered_prompt=answered_prompt,
                )
            )
            yes_no_markers = ("是否", "有没有", "能否", "亮", "正常", "一致")
            result["diagnosis"]["pending_question"] = pending_question
            result["diagnosis"]["reply_options"] = (
                ["是", "否", "不确定"]
                if pending_question
                and any(marker in pending_question for marker in yes_no_markers)
                else []
            )
            result["diagnosis"]["context_applied"] = bool(
                reply_context.get("applied")
            )
            result["conversation_context"] = {
                "continued": bool(reply_context.get("applied")),
                "answered_prompt": answered_prompt,
                "pending_question": pending_question,
                "topic": diagnostic_topic,
            }
            if image_ocr_text:
                result["image_recognition"] = {
                    "used": True,
                    "text": image_ocr_text,
                    "file_name": str(payload.get("image_name", "图片")),
                }
            if include_images:
                schedule_preview_warmup(result.get("sources", []))

            response_persistence_started = time.perf_counter()
            with connect() as connection:
                history_content = str(result["answer"])
                if pending_question:
                    history_content += f"\n\n【待确认问题】{pending_question}"
                cursor = connection.execute(
                    """
                    INSERT INTO messages
                    (conversation_id, role, content, citations_json, created_at)
                    VALUES (?, 'assistant', ?, ?, ?)
                    """,
                    (
                        conversation_id,
                        history_content,
                        json.dumps(result["sources"], ensure_ascii=False),
                        utc_now(),
                    ),
                )
                connection.execute(
                    """
                    UPDATE conversations
                    SET pending_question = ?, diagnostic_topic = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        pending_question,
                        diagnostic_topic,
                        utc_now(),
                        conversation_id,
                    ),
                )
                connection.commit()
                message_id = cursor.lastrowid
            persistence_ms += round(
                (time.perf_counter() - response_persistence_started) * 1000,
                2,
            )
            result["conversation_id"] = conversation_id
            result["message_id"] = message_id
            total_ms = round((time.perf_counter() - request_started) * 1000, 2)
            result["timing"] = {
                "retrieval_ms": retrieval_ms,
                "first_token_ms": first_token_ms,
                "agent_ms": agent_ms,
                "persistence_ms": round(persistence_ms, 2),
                "total_ms": total_ms,
                "cache_hit": cache_hit,
            }
            record_performance(
                {
                    "retrieval_ms": retrieval_ms,
                    "agent_ms": agent_ms,
                    "total_ms": total_ms,
                    "cache_hit": cache_hit,
                    "use_agent": True,
                    "answer_mode": answer_mode,
                    "time": utc_now(),
                }
            )
            self.send_stream_event({"type": "done", "data": result})
        except (BrokenPipeError, ConnectionResetError):
            return
        except Exception as exc:
            try:
                self.send_stream_event(
                    {
                        "type": "error",
                        "error": f"{type(exc).__name__}: {exc}",
                    }
                )
            except (BrokenPipeError, ConnectionResetError):
                return
        finally:
            try:
                self.finish_stream()
            except (BrokenPipeError, ConnectionResetError):
                pass

    def handle_image_recognition(self, payload: dict[str, object]) -> None:
        encoded = str(payload.get("image_base64", "")).strip()
        if not encoded:
            raise ValueError("请选择需要识别的图片")
        if "," in encoded and encoded.lower().startswith("data:image/"):
            encoded = encoded.split(",", 1)[1]
        try:
            data = base64.b64decode(encoded, validate=True)
        except (ValueError, binascii.Error) as exc:
            raise ValueError("图片编码无效") from exc
        file_name = str(payload.get("file_name", "图片.jpg")).strip() or "图片.jpg"
        suffix = Path(file_name).suffix.lower() or ".jpg"
        result = recognize_image_bytes(data, suffix=suffix)
        text = str(result.get("text", "")).strip()
        if not text:
            raise ValueError("未从图片中识别到清晰文字，请重拍或手动输入故障现象")
        result["ok"] = True
        result["file_name"] = file_name
        result["suggested_question"] = (
            "请结合知识库分析图片中的故障码和故障现象。"
        )
        self.send_json(result)

    def handle_feedback(self, payload: dict[str, object]) -> None:
        rating = str(payload.get("rating", "")).strip()
        if rating not in {"up", "down"}:
            raise ValueError("rating 只能是 up 或 down")
        message_id = int(payload.get("message_id") or 0)
        with connect() as connection:
            if message_id:
                exists = connection.execute(
                    "SELECT 1 FROM feedback WHERE message_id = ?",
                    (message_id,),
                ).fetchone()
                if exists:
                    raise ValueError("该回答已经评价，不能修改")
            connection.execute(
                """
                INSERT INTO feedback
                (conversation_id, message_id, account, vehicle_id, question,
                 answer, rating, comment, vehicle_series, scene, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    str(payload.get("conversation_id", "")),
                    message_id or None,
                    str(payload.get("account", "")),
                    str(payload.get("vehicle_id", "")),
                    str(payload.get("question", "")),
                    str(payload.get("answer", "")),
                    rating,
                    str(payload.get("comment", "")).strip()[:1000],
                    str(payload.get("vehicle_series", "")),
                    str(payload.get("scene", "")),
                    utc_now(),
                ),
            )
            connection.commit()
        self.send_json({"saved": True, "rating": rating})

    def handle_intent(self, payload: dict[str, object]) -> None:
        intent_id = re.sub(r"[^a-z0-9_-]", "", str(payload.get("id", "")).lower())
        name = str(payload.get("name", "")).strip()
        if not intent_id or not name:
            raise ValueError("意图ID和名称不能为空")
        scene = str(payload.get("scene", "")).strip()
        if scene not in {"", "用", "养", "修", "保"}:
            raise ValueError("scene 只能是用、养、修、保或空")
        with connect() as connection:
            connection.execute(
                """
                INSERT INTO query_intents
                (id, name, scene, description, example_question, sort_order, enabled)
                VALUES (?, ?, ?, ?, ?, ?, 1)
                ON CONFLICT(id) DO UPDATE SET name=excluded.name, scene=excluded.scene,
                  description=excluded.description,
                  example_question=excluded.example_question,
                  sort_order=excluded.sort_order, enabled=1
                """,
                (
                    intent_id,
                    name,
                    scene,
                    str(payload.get("description", "")).strip(),
                    str(payload.get("example_question", "")).strip(),
                    int(payload.get("sort_order") or 100),
                ),
            )
            connection.commit()
        self.send_json({"saved": True, "id": intent_id})

    def handle_vin_record(self, payload: dict[str, object]) -> None:
        vin = str(payload.get("vin", "")).strip().upper()
        if not re.fullmatch(r"[A-HJ-NPR-Z0-9]{17}", vin):
            raise ValueError("请输入正确的17位VIN")
        fields = (
            "vehicle_type", "chassis_no", "emission_type", "vehicle_series", "fuel_type",
            "announcement_model", "factory_model_code", "rear_axle", "tire_spec",
            "engine_type", "engine_model", "transmission_model", "offline_time",
            "vehicle_note", "engine_name", "device_app_version", "mcu_version",
            "sim_match",
        )
        values = [str(payload.get(field, "")).strip() for field in fields]
        with connect() as connection:
            connection.execute(
                f"""
                INSERT INTO vin_records (vin, {', '.join(fields)}, updated_at)
                VALUES ({', '.join('?' for _ in range(len(fields) + 2))})
                ON CONFLICT(vin) DO UPDATE SET
                {', '.join(f'{field}=excluded.{field}' for field in fields)},
                updated_at=excluded.updated_at
                """,
                (vin, *values, utc_now()),
            )
            connection.commit()
        self.send_json({"saved": True, "vin": vin})

    def send_local_file(self, target: Path, inline: bool = True) -> None:
        data = target.read_bytes()
        content_type = mimetypes.guess_type(target.name)[0] or "application/octet-stream"
        disposition = "inline" if inline else "attachment"
        encoded_name = quote(target.name)
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.send_header(
            "Content-Disposition",
            f"{disposition}; filename*=UTF-8''{encoded_name}",
        )
        self.send_header("Cache-Control", "private, max-age=3600")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.end_headers()
        self.wfile.write(data)

    def serve_static(self, request_path: str) -> None:
        if request_path in {"", "/"}:
            relative = "index.html"
        elif request_path.endswith("/"):
            relative = request_path.lstrip("/") + "index.html"
        else:
            relative = request_path.lstrip("/")
        target = (WEB_DIR / relative).resolve()
        if WEB_DIR.resolve() not in target.parents and target != WEB_DIR.resolve():
            self.send_error(HTTPStatus.FORBIDDEN)
            return
        if not target.exists() or not target.is_file():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        data = target.read_bytes()
        content_type = mimetypes.guess_type(target.name)[0] or "application/octet-stream"
        if content_type.startswith("text/") or content_type in {
            "application/javascript",
            "application/json",
        }:
            content_type += "; charset=utf-8"
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(data)


def main() -> int:
    if not DATABASE_PATH.exists():
        print("知识库尚未构建，请先运行 scripts/build_kb.py", file=sys.stderr)
        return 2
    ensure_runtime_schema()
    warm_retrieval_runtime()
    server = ThreadingHTTPServer((HOST, PORT), KnowledgeHandler)
    print(f"知识库服务已启动：http://{HOST}:{PORT}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
